import openpyxl
import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import re
import os
from prettytable import PrettyTable
import sys
import sqlite3
from sqlite3 import Error
import pandas as pd

try:
    with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute('CREATE TABLE IF NOT EXISTS clientes (ClaveCliente INTEGER PRIMARY KEY, nombre TEXT NOT NULL, correo TEXT NOT NULL, rfc TEXT NOT NULL, estadoC BLOB NOT NULL);')
        mi_cursor.execute('CREATE TABLE IF NOT EXISTS servicios (ClaveServicio INTEGER PRIMARY KEY, descripcion TEXT NOT NULL, costo FLOAT NOT NULL, estadoS BLOB NOT NULL);')
        mi_cursor.execute('CREATE TABLE IF NOT EXISTS notas (folio INTEGER PRIMARY KEY, fecha timestamp NOT NULL, ClaveCliente INTEGER, monto REAL NOT NULL,estadoN BLOB NOT NULL ,FOREIGN KEY (ClaveCliente) REFERENCES clientes (ClaveCliente));')
        mi_cursor.execute('CREATE TABLE IF NOT EXISTS detalle (folio INTEGER, ClaveServicio INTEGER, FOREIGN KEY (folio) REFERENCES notas (folio), FOREIGN KEY (ClaveServicio) REFERENCES servicios (ClaveServicio));')
        print('TABLAS CREADAS EXITOSAMENTE')
        servicios = [
        ("Cambio de aceite y filtro", 1600.0, 1),
        ("Limpieza del filtro de aire, de gasolina y las bujías", 4000.0, 1),
        ("Afinacion menor", 1050.0, 1),
        ("Afinacion mayor", 2200.0, 1),
        ("Revisión de motor", 5000.0, 1)
        ]
        for servicio in servicios:
          mi_cursor.execute('INSERT INTO servicios (descripcion, costo, estadoS) VALUES (?, ?, ?)', servicio)

except Error as e:
    print(e)
except Exception:
    print(f'Se produjo el siguiente error: {sys.exc_info()[0]}')

def registrar_nota():
    while True:
        print('\n═══════════════════════════════════')
        print('     REGISTRAR UNA NUEVA NOTA')
        print('═══════════════════════════════════')
        monto_total = 0
        servicios_seleccionados = []
        try:
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute('SELECT ClaveCliente, nombre FROM clientes WHERE estadoC = 1')
                datos = mi_cursor.fetchall()
                tabla = PrettyTable()
                tabla.field_names = ["CLAVE", "NOMBRE"]
                print('\n    Clientes Registrados')
                for dato in datos:
                    tabla.add_row(dato)
                print(tabla)
                if not datos:
                    print("NO HAY CLIENTES REGISTRADOS")
                while True:
                    try:
                        cliente = int(input('\nIngrese la clave del cliente o ingrese "0" para regresar al menú anterior: '))
                        if cliente == 0:
                            print('\n** OPERACIÓN CANCELADA. VOLVIENDO AL MENÚ DE NOTAS **')
                            return
                        else:
                            mi_cursor.execute('SELECT ClaveCliente FROM clientes WHERE estadoC = 1 AND ClaveCliente = ?', (cliente,))
                            cliente_existente = mi_cursor.fetchall()
                            if cliente_existente:
                                break
                            else:
                                print('\n** EL CLIENTE NO ESTÁ REGISTRADO O CORRESPONDE A UN CLIENTE SUSPENDIDO. **')
                        continue
                    except ValueError:
                        print(f'\n** DATO NO VÁLIDO. POR FAVOR, INGRESE UN DATO VÁLIDO. **')
                while True:
                    try:
                        fecha_str = input('\nIngrese la fecha en formato DD/MM/YYYY: ')
                        fecha_actual = datetime.now().date()
                        fecha_ingresada = datetime.strptime(fecha_str, "%d/%m/%Y").date()
                        if fecha_ingresada > fecha_actual:
                            print('\n** LA FECHA INGRESADA NO DEBE SER POSTERIOR A LA FECHA ACTUAL. INGRESE UNA FECHA VÁLIDA **')
                        else:
                            break
                    except Exception:
                        print(f'\n** DATO NO VÁLIDO. POR FAVOR, INGRESE LA FECHA EN EL FORMATO CORRECTO. **')
                mi_cursor.execute('SELECT ClaveServicio, descripcion, costo FROM servicios WHERE estadoS = 1')
                datos_s = mi_cursor.fetchall()
                tabla = PrettyTable()
                tabla.field_names = ['CLAVE', 'DESCRIPCIÓN', 'COSTO']
                print('\n      Servicios Registrados')
                for dato in datos_s:
                    tabla.add_row(dato)
                print(tabla)
                while True:
                    try:
                        servicio = int(input('\nIngrese la clave del servicio o ingrese "0" para volver al menú anterior: '))
                        if servicio == 0:
                            return
                        mi_cursor.execute('SELECT ClaveServicio, costo FROM servicios WHERE ClaveServicio = ? AND estadoS = 1', (servicio,))
                        servicio_existente = mi_cursor.fetchone()
                        if servicio_existente:
                            monto_total += servicio_existente[1]
                            servicios_seleccionados.append(int(servicio))
                            agregar_otro_servicio = input('\n¿Deseas agregar otro servicio? (S)i (N)o: ')
                            if agregar_otro_servicio.lower() == 's':
                                continue
                            elif agregar_otro_servicio.strip() == '':
                                print("ERROR, NO SE PUEDE OMITIR, INGRESE UN DATO VÁLIDO")
                            else:
                                mi_cursor.execute('INSERT INTO notas (fecha, ClaveCliente, monto, estadoN) VALUES (?, ?, ?, ?)',
                                                  (fecha_ingresada, cliente, monto_total, 1))
                                id_nota = mi_cursor.lastrowid
                                for id_servicio in servicios_seleccionados:
                                    valores_detalle = (id_nota, id_servicio)
                                    mi_cursor.execute('INSERT INTO detalle (Folio, ClaveServicio) VALUES (?, ?)', valores_detalle)
                                print('\n** NOTA(S) REGISTRADA(S) CORRECTAMENTE **')
                                break
                        else:
                            print('\n** LA CLAVE DE SERVICIO INGRESADA NO ESTÁ REGISTRADA O CORRESPONDE A UN SERVICIO CANCELADO. **')
                            continue
                    except ValueError:
                        print(f'\n** DATO NO VÁLIDO. POR FAVOR, INGRESE UN DATO VÁLIDO. **')
                while True:
                    agregar_nota = input('\n¿Deseas registrar otra nota? (S)i (N)o: ')
                    if agregar_nota.lower() == 'n':
                        return
                    elif agregar_nota.lower() == 's':
                        break
                    else:
                        print('\n** DATO NO VÁLIDO. POR FAVOR, INGRESE (S) PARA CONFIRMAR LA ACCIÓN O (N) PARA CANCELAR LA OPERACIÓN **')
        except Error as e:
            print(e)
        except Exception:
            print(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
        finally:
            conn.close()

def cancelar_nota():
    while True:
        print('\n════════════════════════════')
        print('     CANCELAR UNA NOTA')
        print('════════════════════════════')
        try:
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                try:
                    folio_cancelar = int(input('\nIngrese el folio de la nota a cancelar o ingrese "0" para volver al menú anterior: '))
                    if folio_cancelar == 0:
                        print('\n** OPERACIÓN CANCELADA. VOLVIENDO AL MENÚ DE NOTAS **')
                        return
                    else:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute('SELECT notas.folio, strftime("%d/%m/%Y", notas.fecha) AS fecha_formateada, clientes.ClaveCliente, clientes.nombre AS nombre_cliente, GROUP_CONCAT(servicios.descripcion, ", ") AS servicios, SUM(servicios.costo) AS costo_total FROM notas INNER JOIN clientes ON notas.ClaveCliente = clientes.ClaveCliente INNER JOIN detalle ON notas.folio = detalle.Folio INNER JOIN servicios ON detalle.ClaveServicio = servicios.ClaveServicio WHERE notas.folio = ? AND notas.estadoN = 1 GROUP BY notas.folio, notas.fecha, clientes.ClaveCliente, clientes.nombre;', (folio_cancelar,))
                        nota_existente = mi_cursor.fetchone()
                        if nota_existente:
                            tabla_detalles = PrettyTable(["Folio", "Fecha", "Clave Cliente", "Nombre Cliente", "Descripción de Servicios", "Costo Total"])
                            tabla_detalles.add_row(nota_existente)
                            print('\n    DETALLES DE LA NOTA')
                            print(tabla_detalles)
                        else:
                            print('\n** ERROR, EL FOLIO INGRESADO NO CORRESPONDE A UNA NOTA ACTIVA **')
                            continue
                except ValueError:
                    print(f'\n** DATO NO VÁLIDO. POR FAVOR, INGRESE UN DATO VÁLIDO. **')
                    continue
                while True:
                    confirmacion = input('\n¿Desea cancelar esta nota? (S)i (N)o: ')
                    if confirmacion.lower() == 's':
                        mi_cursor = conn.cursor()
                        mi_cursor.execute('UPDATE notas SET estadoN = 0 WHERE folio = ?', (folio_cancelar,))
                        print('\n** NOTA CANCELADA CORRECTAMENTE **')
                        break
                    elif confirmacion.lower() == 'n':
                        print('\n** OPERACIÓN CANCELADA **')
                        break
                    else:
                        print('\n** DATO NO VÁLIDO. POR FAVOR, INGRESE (S) PARA CONFIRMAR LA ACCIÓN O (N) PARA CANCELAR LA OPERACIÓN **')
        except sqlite3.Error as e:
            print(e)
        except Exception:
            print(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
        finally:
            conn.close()

def recuperar_nota():
    while True:
        print('\n════════════════════════════')
        print('    RECUPERAR UNA NOTA')
        print('════════════════════════════')
        try:
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute('SELECT folio FROM notas WHERE estadoN = 0')
                notas_canceladas = [folio[0] for folio in mi_cursor.fetchall()]
                if not notas_canceladas:
                    print('*  No hay notas canceladas para recuperar *')
                    return
                print('\n** NOTAS PREVIAMENTE CANCELADAS: **')
                print('---------------------------------------')
                prettytable1 = PrettyTable()
                prettytable1.field_names = ["Folio"]
                for folio in notas_canceladas:
                    prettytable1.add_row([folio])
                print(prettytable1)
                while True:
                    try:
                        folio_recuperar = int(input('\nIngrese el folio de la nota a recuperar o ingrese "0" para regresar al menú anterior: '))
                        if folio_recuperar == 0:
                            print('\n** OPERACIÓN CANCELADA. VOLVIENDO AL MENÚ DE NOTAS **')
                            return
                        elif folio_recuperar in notas_canceladas:
                            mi_cursor.execute('SELECT notas.folio, strftime("%d/%m/%Y", notas.fecha) AS fecha_formateada, clientes.ClaveCliente, clientes.nombre AS nombre_cliente, GROUP_CONCAT(servicios.descripcion, ", ") AS servicios, SUM(servicios.costo) AS costo_total FROM notas INNER JOIN clientes ON notas.ClaveCliente = clientes.ClaveCliente INNER JOIN detalle ON notas.folio = detalle.Folio INNER JOIN servicios ON detalle.ClaveServicio = servicios.ClaveServicio WHERE notas.folio = ? AND notas.estadoN = 0 GROUP BY notas.folio, notas.fecha, clientes.ClaveCliente, clientes.nombre;', (folio_recuperar,))
                            detalle_nota = mi_cursor.fetchone()
                            tabla_detalles = PrettyTable(["Folio", "Fecha", "Clave Cliente", "Nombre Cliente", "Descripción de Servicios", "Costo Total"])
                            tabla_detalles.add_row(detalle_nota)
                            print('\n    DETALLES DE LA NOTA A RECUPERAR')
                            print(tabla_detalles)
                            while True:
                                confirmacion = input('\n¿Desea recuperar esta nota? (S)i (N)o: ')
                                if confirmacion.lower() == 's':
                                    mi_cursor.execute('UPDATE notas SET estadoN = 1 WHERE folio = ?', (folio_recuperar,))
                                    print(f'\n** Nota {folio_recuperar} recuperada **')
                                    break
                                elif confirmacion.lower() == 'n':
                                    print('\n** OPERACION CANCELADA **')
                                    break
                                else:
                                    print('\n** DATO NO VÁLIDO. POR FAVOR, INGRESE (S) PARA CONFIRMAR LA ACCIÓN O (N) PARA CANCELAR LA OPERACIÓN **')
                                    continue
                        else:
                            print('\n** El FOLIO INGRESADO NO CORRESPONDE A UNA NOTA CANCELADA **')
                    except ValueError:
                        print(f'\n** DATO NO VÁLIDO. POR FAVOR, INGRESE UN DATO VÁLIDO. **')
        except sqlite3.Error as e:
            print(e)
        except Exception:
            print(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
        finally:
            conn.close()

def consultas_reportes_notas():
    while True:
        print('\n═══════════════════════════════════')
        print('    CONSULTAS Y REPORTES NOTAS')
        print('═══════════════════════════════════')
        print('1. Consulta por período')
        print('2. Consulta por folio')
        print('3. Volver al menú notas')
        try:
            consulta = int(input('\nIngresa el número de la operación que desea realizar: '))
            if 1 <= consulta <= 3:
                if consulta == 1:
                    try:
                      with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                        mi_cursor = conn.cursor()
                        fecha_inicial_str = input('\nIngrese la fecha inicial (DD/MM/YYYY) o presione Enter para usar fecha dada por el sistema (01/01/2000): ')
                        if fecha_inicial_str == '':
                            fecha_inicial = datetime(2000, 1, 1).date()
                        else:
                          try:
                            fecha_inicial = datetime.strptime(fecha_inicial_str, "%d/%m/%Y").date()
                          except ValueError:
                            print('\n** FORMATO DE FECHA INVÁLIDO. Intente de nuevo. **')
                            continue
                        while True:
                          fecha_final_str = input('\nIngrese la fecha final (DD/MM/YYYY) o presione Enter para usar la fecha actual: ')
                          if fecha_final_str == '':
                            fecha_final = datetime.today().date()
                          else:
                            try:
                               fecha_final = datetime.strptime(fecha_final_str, "%d/%m/%Y").date()
                            except ValueError:
                               print('\n** FORMATO DE FECHA INVÁLIDO. Intente de nuevo. **')
                               continue
                          if fecha_final < fecha_inicial:
                            print('\n** LA FECHA FINAL DEBE SER MAYOR O IGUAL QUE LA FECHA ACTUAL. **')
                            continue
                          else:
                            break
                        mi_cursor.execute('SELECT folio, fecha, ClaveCliente, monto FROM notas WHERE fecha BETWEEN ? AND ? AND estadoN = True;', (fecha_inicial, fecha_final))
                        registros = mi_cursor.fetchall()
                        if registros:
                          columnas = ['FOLIO', 'FECHA', 'CLAVE CLIENTE', 'MONTO']
                          df = pd.DataFrame(registros, columns=columnas)
                          monto_promedio = df['MONTO'].mean()
                          tabla = PrettyTable(columnas)
                          for registro in registros:
                              tabla.add_row(registro)
                          print(tabla)    
                          print('\nMonto promedio en el periodo: ${:.2f}'.format(monto_promedio))
                        else:
                          print('\nNo se encontraron registros en el periodo seleccionado.')
                          print(tabla)
                          print('\n---------------------------------------')
                          print('           EXPORTAR REPORTE')
                          print('---------------------------------------')
                          print('1. Exportar reporte como archivo EXCEL')
                          print('2. Exportar reporte como archivo CSV')
                          print('3. Volver al menú consultas y reportes')
                          exportar = int(input('\nIngresa el número de la operación que deseas realizar: '))
                          encabezados = ['FOLIO', 'FECHA', 'CLAVE CLIENTE', 'MONTO']
                          if exportar == 1:
                             nombre_excel = f'ReportePorPeriodo_{fecha_inicial}_{fecha_final}.xlsx'
                             wb = Workbook()
                             hoja = wb.active
                             hoja.append(encabezados)
                             for dato in datos:
                                hoja.append(dato)
                                wb.save(nombre_excel)
                                print(f'\nInforme {nombre_excel} exportado correctamente')
                          elif exportar == 2:
                                nombre_csv = f'ReportePorPeriodo_{fecha_inicial}_{fecha_final}.csv'
                                with open(nombre_csv, 'w', newline='') as reporte_csv:
                                  grabador = csv.writer(reporte_csv)
                                  grabador.writerow(encabezados)
                                  grabador.writerows(datos)
                                  print(f'\nInforme {nombre_csv} exportado correctamente')
                          elif exportar == 3:
                              continue
                          else:
                              print(f'\n** NO HAY NOTAS EMITIDAS EN EL PERIODO {fecha_inicial} A {fecha_final} **')
                    except sqlite3.Error as e:
                           print(e)
                    except Exception:
                           print(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
                    finally:
                        conn.close()
                elif consulta == 2:
                    try:
                        with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                            conn.row_factory = sqlite3.Row
                            mi_cursor = conn.cursor()
                            mi_cursor.execute('SELECT n.folio, n.fecha, c.nombre FROM notas n, clientes c WHERE n.ClaveCliente = c.ClaveCliente AND estadoN = 1 ORDER BY n.folio')
                            datos = mi_cursor.fetchall()
                            tabla = PrettyTable()
                            tabla.field_names = ['FOLIO', 'FECHA', 'NOMBRE CLIENTE']
                            print('')
                            for dato in datos:
                                tabla.add_row(dato)
                            print(tabla)
                            while True:
                                clave_consultar = input('\nIngrese el folio de la nota que desea consultar o ingrese "0" para volver): ')
                                if clave_consultar == '0':
                                    break
                                datos = {'folio': int(clave_consultar)}
                                mi_cursor.execute('''SELECT notas.folio,
                                               notas.fecha,
                                               clientes.ClaveCliente,
                                               clientes.nombre,
                                               clientes.correo,
                                               clientes.rfc,
                                               GROUP_CONCAT(servicios.descripcion || ' - $' || CAST(servicios.costo AS TEXT), '\n') as descripcion_costo,
                                               SUM(servicios.costo)
                                            FROM notas
                                            INNER JOIN clientes ON notas.ClaveCliente = clientes.ClaveCliente
                                            INNER JOIN detalle ON notas.folio = detalle.Folio
                                            INNER JOIN servicios ON detalle.ClaveServicio = servicios.ClaveServicio
                                            WHERE notas.folio = :folio AND notas.estadoN = True
                                            GROUP BY notas.folio, notas.fecha, clientes.ClaveCliente, clientes.nombre;''', datos)
                                registro = mi_cursor.fetchall()
                                if registro:
                                    nueva_tabla = PrettyTable()
                                    nueva_tabla.field_names = ["Folio", "Fecha", "Clave Cliente", "Nombre Cliente", "Correo Cliente", "RFC cliente", "Descripción de Servicios", "Costo Total"]
                                    print('')
                                    for fila in registro:
                                        nueva_tabla.add_row(fila)
                                    print(nueva_tabla)
                                else:
                                    print(f'\n** LA NOTA ASOCIADA AL FOLIO {clave_consultar} NO HA SIDO ENCONTRADA O SE ENCUENTRA CANCELADA. POR FAVOR, VERIFICA LA INFORMACIÓN E INTENTA NUEVAMENTE. **')
                    except sqlite3.Error as e:
                        print(e)
                    except Exception:
                        print(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
                    finally:
                        conn.close()
                elif consulta == 3:
                    print('\n** OPERACIÓN CANCELADA. VOLVIENDO AL MENÚ DE NOTAS **')
                    return
            else:
                raise ValueError
        except ValueError:
            print('\n** DATO NO VÁLIDO. INGRESE EL NÚMERO DE ALGUNA OPCIÓN MOSTRADA **')

def menu_notas():
    while True:
        print('\n---------------------------------------')
        print('      BIENVENIDO AL MENU DE NOTAS   ')
        print('---------------------------------------')
        print('1. Registrar una nota')
        print('2. Cancelar una nota')
        print('3. Recuperar una nota')
        print('4. Consultas y reportes')
        print('5. Salir')
        opcion = input('\nIngrese el número de la opción a la que desea ingresar: ')
        if opcion.isdigit():
            opcion = int(opcion)
            if opcion == 1:
                registrar_nota()
            elif opcion == 2:
                cancelar_nota()
            elif opcion == 3:
                recuperar_nota()
            elif opcion == 4:
                consultas_reportes_notas()
            elif opcion == 5:
              menu_principal()
            else:
                print('\nERROR, POR FAVOR INGRESE UNA OPCIÓN VÁLIDA.')
        else:
            print('\nERROR, POR FAVOR INGRESE UNA OPCIÓN VÁLIDA.')
            
def agregar_cliente():
    while True:
        print('\n════════════════════════════')
        print('  REGISTRA UN NUEVO CLIENTE')
        print('════════════════════════════')
        print('\nNota: Ingrese "0" para volver al menú anterior')
        estadoC = True
        while True:
            try:
                nombre = input('\nIngrese el nombre completo del cliente: ')
                if nombre == '0':
                    print('\n** OPERACION CANCELADA. VOLVIENDO AL MENÚ CLIENTES **')
                    return
                if nombre.isdigit():
                   print('\n** DATO NO VÁLIDO. INGRESE UN NOMBRE VÁLIDO **')
                   continue
                if nombre.strip() == '':
                    print('\n** EL DATO NO PUEDE OMITIRSE. INGRESE UN DATO O (0) PARA VOLVER AL MENÚ CLIENTES **')
                    continue
                else:
                    break
            except ValueError as e:
                print(e)
        while True:
            try:
                correo = input('\nIngrese el correo electrónico del cliente: ')
                if correo == '0':
                    print('\n** OPERACION CANCELADA. VOLVIENDO AL MENÚ CLIENTES **')
                    return
                if correo.strip() == '':
                    print('\n** EL DATO NO PUEDE OMITIRSE. INGRESE UN DATO O (0) PARA VOLVER AL MENÚ CLIENTES **')
                    continue
                elif validar_correo(correo):
                    break
                else:
                    print('\n** EL CORREO PROPORCIONADO NO TIENE UN FORMATO VÁLIDO. INGRESELO NUEVAMENTE **')
                    continue
            except Error as e:
                print(e)
        while True:
            try:
                rfc = input('\nIngrese el rfc del cliente: ')
                if rfc == '0':
                    print('\n** OPERACION CANCELADA. VOLVIENDO AL MENÚ CLIENTES **')
                    return
                if rfc.strip() == '':
                    print('\n** EL DATO NO PUEDE OMITIRSE. INGRESE UN DATO VÁLIDO **')
                    continue
                elif validar_rfc(rfc):
                    break
                else:
                    print('\n** EL RFC PROPORCIONADO NO TIENE UN FORMATO VÁLIDO. INGRESELO NUEVAMENTE **')
                    continue
            except Error as e:
                print(e)
        try:
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                datos = (nombre, correo, rfc, estadoC)
                mi_cursor.execute('INSERT INTO clientes (nombre,correo,rfc,estadoC) VALUES (?,?,?,?)', datos)
                print('\n** Cliente registrado correctamente **')
                print(f'Clave asignada: {mi_cursor.lastrowid}')
        except Error as e:
            print(f'SE PRODUJO EL SIGUIENTE ERROR: {e}')
        finally:
            conn.close()
        while True:
            agregar_cliente = input('\n¿Desea agregar otro cliente? (Si) (No): ')
            if agregar_cliente.lower() == 's':
                break
            elif agregar_cliente.lower() == 'n':
                print('\n** OPERACIÓN CANCELADA. VOLVIENDO AL MENÚ CLIENTES**')
                return
            else:
                print('\n** DATO NO VÁLIDO. INGRESE (S)I Ó (N)O **')
                continue

def suspender_cliente():
    while True:
        print('\n════════════════════════════')
        print('   SUSPENDER UN CLIENTE')
        print('════════════════════════════')
        print('\nNota: Escriba (0) si desea volver al menú clientes')
        try:
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute('SELECT ClaveCliente, nombre FROM clientes WHERE estadoC = True')
                datos = mi_cursor.fetchall()
                tabla = PrettyTable()
                tabla.field_names = ["CLAVE", "NOMBRE"]
                print('\n    Clientes Registrados')
                for dato in datos:
                    tabla.add_row(dato)
                print(tabla)
            while True:
                try:
                    clave_suspender = int(input('\nIngrese la clave del cliente a suspender: '))
                    if clave_suspender == 0:
                        print('\n** OPERACION CANCELADA. VOLVIENDO AL MENÚ CLIENTES **')
                        return
                    mi_cursor = conn.cursor()
                    mi_cursor.execute('SELECT ClaveCliente FROM clientes WHERE ClaveCliente = ? AND estadoC = True', (clave_suspender,))
                    cliente_existente = mi_cursor.fetchall()
                    if cliente_existente:
                      mi_cursor.execute('SELECT ClaveCliente, nombre, correo, rfc FROM clientes WHERE ClaveCliente = ?', (clave_suspender,))
                      datos = mi_cursor.fetchall()
                      tabla = PrettyTable()
                      tabla.field_names = ['CLAVE', 'NOMBRE','CORREO','RFC']
                      print('\n    Cliente a suspender')
                      for dato in datos:
                        tabla.add_row(dato)
                      print(tabla)
                      while True:
                        confirmacion = input('\n¿Está seguro que desea suspender este registro? (S)i (N)o: ')
                        if confirmacion.lower() == 's':
                            mi_cursor.execute('UPDATE clientes SET estadoC = False WHERE ClaveCliente = ?', (clave_suspender,))
                            conn.commit()
                            print('\n** CLIENTE SUSPENDIDO CORRECTAMENTE **')
                            return
                        elif confirmacion.lower() == 'n':
                            print('\n** OPERACION CANCELADA. VOLVIENDO AL MENÚ CLIENTES **')
                            return
                        else:
                            print('\n** DATO NO VÁLIDO. POR FAVOR, INGRESE (S) PARA CONFIRMAR LA ACCIÓN O (N) PARA CANCELAR LA OPERACIÓN **')
                    else:
                      print('\n** EL CLIENTE NO EXISTE O SE ENCUENTRA SUSPENDIDO. **')
                      continue
                except Exception:
                  print(f'\n** DATO NO VÁLIDO. INGRESE LA CLAVE DE ALGÚN CLIENTE **')
        except Error as e:
            print(e)
        except Exception:
            print(f'Se produjo el siguiente error: {sys.exc_info()[0]}')          
        finally:
            conn.close()

def recuperar_cliente():
    while True:
        print('\n════════════════════════════')
        print('   RECUPERAR UN CLIENTE')
        print('════════════════════════════')
        print('\nNota: Escriba (0) si desea volver al menú clientes')
        try:
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute('SELECT ClaveCliente, nombre FROM clientes WHERE estadoC = False')
                datos = mi_cursor.fetchall()
                tabla = PrettyTable()
                tabla.field_names = ["CLAVE", "NOMBRE"]
                print('\n Clientes Suspendidos')
                for dato in datos:
                    tabla.add_row(dato)
                print(tabla)
            while True:
                try:
                    clave_recuperar = int(input('\nIngrese la clave del cliente a recuperar: '))
                    if clave_recuperar == 0:
                        print('\n** OPERACION CANCELADA. VOLVIENDO AL MENÚ CLIENTES **')
                        return
                    mi_cursor = conn.cursor()
                    mi_cursor.execute('SELECT ClaveCliente FROM clientes WHERE ClaveCliente = ? AND estadoC = False', (clave_recuperar,))
                    cliente_suspendido = mi_cursor.fetchall()
                    if cliente_suspendido:
                        mi_cursor.execute('SELECT ClaveCliente, nombre, correo, rfc FROM clientes WHERE ClaveCliente = ?', (clave_recuperar,))
                        datos = mi_cursor.fetchall()
                        tabla = PrettyTable()
                        tabla.field_names = ['CLAVE', 'NOMBRE','CORREO','RFC']
                        print('\n    Cliente a recuperar')
                        for dato in datos:
                            tabla.add_row(dato)
                        print(tabla)
                        while True:
                            confirmacion = input('\n¿Está seguro que desea recuperar este registro? (S)i (N)o: ')
                            if confirmacion.lower() == 's':
                                mi_cursor.execute('UPDATE clientes SET estadoC = True WHERE ClaveCliente = ?', (clave_recuperar,))
                                conn.commit()
                                print('\n** ClIENTE RECUPERADO CORRECTAMENTE **')
                                return
                            elif confirmacion.lower() == 'n':
                                print('\n** OPERACION CANCELADA. VOLVIENDO AL MENÚ CLIENTES **')
                                return
                            else:
                                print('\n** DATO NO VÁLIDO. POR FAVOR, INGRESE (S) PARA CONFIRMAR LA ACCIÓN O (N) PARA CANCELAR LA OPERACIÓN **')
                    else:
                        print('\n** EL CLIENTE NO SE ENCUENTRA SUSPENDIDO. **')
                except Exception:
                    print(f'\n** DATO NO VÁLIDO. INGRESE LA CLAVE DE ALGÚN CLIENTE **')
        except Error as e:
            print(e)
        except Exception:
            print(f'Se produjo el siguiente error: {sys.exc_info()[0]}')          
        finally:
            conn.close()

def listado_clientes_registrados():
    while True:
        print('\n-----------------------------------------')
        print(' SUBMENÚ LISTADO DE CLIENTES REGISTRADOS')
        print('-----------------------------------------')
        print('1. Ordenado por clave')
        print('2. Ordenado por nombre')
        print('3. Volver al menú anterior')
        opcion = int(input('\nIngresa el número de la operación que deseas realizar: '))
        encabezados = ['CLAVE', 'NOMBRE', 'CORREO', 'RFC']
        if opcion == 1:
            try:
                with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                    mi_cursor = conn.cursor()
                    mi_cursor.execute('SELECT ClaveCliente, nombre, correo, rfc FROM clientes WHERE estadoC = True ORDER BY ClaveCliente')
                    datos = mi_cursor.fetchall()
                    tabla = PrettyTable()
                    tabla.field_names = (encabezados)
                    print(f'\n                    CLIENTES ACTIVOS')
                    for dato in datos:
                        tabla.add_row(dato)
                    print(tabla)
                    print('\n---------------------------------------')
                    print('           EXPORTAR REPORTE')
                    print('---------------------------------------')
                    print('1. Exportar reporte como archivo EXCEL')
                    print('2. Exportar reporte como archivo CSV')
                    print('3. Volver al menú de reportes')
                    exportar = int(input('\nIngresa el número de la operación que deseas realizar: '))
                    if exportar == 1:
                        fecha_reporte = datetime.now().strftime('%d_%m_%Y')
                        nombre_excel = f'ReporteClientesActivosPorClave_{fecha_reporte}.xlsx'
                        wb = Workbook()
                        hoja = wb.active
                        hoja.append(encabezados)
                        for dato in datos:
                            hoja.append(dato)
                        wb.save(nombre_excel)
                        print(f'\nInforme {nombre_excel} exportado correctamente')
                    elif exportar == 2:
                        fecha_reporte = datetime.now().strftime('%d_%m_%Y')
                        nombre_csv = f'ReporteClientesActivosPorClave_{fecha_reporte}.csv'
                        with open(nombre_csv, 'w', newline='') as reporte_csv:
                            grabador = csv.writer(reporte_csv)
                            grabador.writerow(encabezados)
                            grabador.writerows(datos)
                        print(f'\nInforme {nombre_csv} exportado correctamente')
                    elif exportar == 3:
                        break
            except Exception as e:
                print(f"Error: {e}")
        elif opcion == 2:
            try:
                with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                    mi_cursor = conn.cursor()
                    mi_cursor.execute('SELECT ClaveCliente, nombre, correo, rfc FROM clientes WHERE estadoC = True ORDER BY nombre')
                    datos = mi_cursor.fetchall()
                    tabla = PrettyTable()
                    tabla.field_names = (encabezados)
                    print(f'\n                    CLIENTES ACTIVOS')
                    for dato in datos:
                        tabla.add_row(dato)
                    print(tabla)
                    print('\n---------------------------------------')
                    print('           EXPORTAR REPORTE')
                    print('---------------------------------------')
                    print('1. Exportar reporte como archivo EXCEL')
                    print('2. Exportar reporte como archivo CSV')
                    print('3. Volver al menú de reportes')
                    exportar = int(input('\nIngresa el número de la operación que deseas realizar: '))
                    if exportar == 1:
                        fecha_reporte = datetime.now().strftime('%d_%m_%Y')
                        nombre_excel = f'ReporteClientesActivosPorNombre_{fecha_reporte}.xlsx'
                        wb = Workbook()
                        hoja = wb.active
                        hoja.append(encabezados)
                        for dato in datos:
                            hoja.append(dato)
                        wb.save(nombre_excel)
                        print(f'\nInforme {nombre_excel} exportado correctamente')
                    elif exportar == 2:
                        fecha_reporte = datetime.now().strftime('%d_%m_%Y')
                        nombre_csv = f'ReporteClientesActivosPorNombre_{fecha_reporte}.csv'
                        with open(nombre_csv, 'w', newline='') as reporte_csv:
                            grabador = csv.writer(reporte_csv)
                            grabador.writerow(encabezados)
                            grabador.writerows(datos)
                        print(f'\nInforme {nombre_csv} exportado correctamente')
                    elif exportar == 3:
                        break
            except Exception as e:
                print(f"Error: {e}")
        elif opcion == 3:
            break
            
def consultas_reportes_clientes():
    while True:
        print('\n═══════════════════════════════════')
        print(' CONSULTAS Y REPORTES DE CLIENTES')
        print('═══════════════════════════════════')
        print('1. Listado de clientes registrados')
        print('2. Búsqueda por clave')
        print('3. Búsqueda por nombre')
        print('4. Volver al menú clientes')        
        try:
            consulta = int(input('\nIngresa el número de la operación que desea realizar: '))          
            if 1 <= consulta <= 4:
                if consulta == 1:
                    listado_clientes_registrados()                   
                elif consulta == 2:
                    try:
                        with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                            mi_cursor = conn.cursor()
                            mi_cursor.execute('SELECT ClaveCliente, nombre FROM clientes ORDER BY ClaveCliente')
                            datos_clientes = mi_cursor.fetchall()
                            tabla_clientes = PrettyTable()
                            tabla_clientes.field_names = ['CLAVE', 'NOMBRE']
                            for cliente in datos_clientes:
                                tabla_clientes.add_row(cliente)                                
                            print(tabla_clientes)                           
                            while True:    
                                clave_consultar = int(input('\nIngrese la clave del cliente a consultar o ingrese "0" para volver al menú anterior: '))                                
                                if clave_consultar == 0:
                                    break                                  
                                mi_cursor.execute('SELECT ClaveCliente, nombre, correo, rfc FROM clientes WHERE ClaveCliente = ?', (clave_consultar,))
                                datos_cliente_consultado = mi_cursor.fetchall()
                                tabla_cliente_consultado = PrettyTable()
                                tabla_cliente_consultado.field_names = ['CLAVE', 'NOMBRE', 'CORREO', 'RFC']                               
                                if datos_cliente_consultado:
                                    print(f'\n    Cliente Encontrado')                                    
                                    for dato in datos_cliente_consultado:
                                        tabla_cliente_consultado.add_row(dato)
                                    print(tabla_cliente_consultado)
                                    break
                                else:
                                    print(f'\n** NO SE ENCONTRÓ UN REGISTRO ASOCIADO A LA CLAVE {clave_consultar}')
                                    break
                    except Exception:
                        print(f'\n** DATO NO VÁLIDO. INGRESE LA CLAVE DE ALGÚN CLIENTE **')                      
                elif consulta == 3:
                    while True:
                        try:
                            nombre_consultar = input('\nIngrese el nombre completo del cliente a consultar o ingrese "0" para volver al menú anterior: ')                           
                            if nombre_consultar == '0':
                                break                               
                            else:
                                with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                                    mi_cursor = conn.cursor()
                                    datos = {'nombre': nombre_consultar.lower()}
                                    mi_cursor.execute('SELECT ClaveCliente, nombre, correo, rfc FROM clientes WHERE LOWER (nombre) = :nombre', datos)
                                    datos1 = mi_cursor.fetchall()
                                    tabla = PrettyTable()
                                    tabla.field_names = ['CLAVE', 'NOMBRE', 'CORREO', 'RFC']                                    
                                    if datos1:
                                        print(f'\n    Cliente Encontrado')                                        
                                        for dato in datos1:
                                            tabla.add_row(dato)                                           
                                        print(tabla)
                                        break                                        
                                    else:
                                        print(f'\n** NO SE ENCONTRÓ UN REGISTRO ASOCIADO {nombre_consultar}')                                        
                        except Exception:
                            print(f'\n** DATO NO VÁLIDO. INGRESE EL NOMBRE DE ALGÚN CLIENTE **')                            
                elif consulta == 4:
                    return
            else:
                print('\n** DATO NO VÁLIDO. INGRESE EL NÚMERO DE ALGUNA OPCIÓN MOSTRADA **')
        except ValueError:
            print('\n** DATO NO VÁLIDO. INGRESE EL NÚMERO DE ALGUNA OPCIÓN MOSTRADA **')

def menu_clientes():
    while True:
        print('\n--------------------------')
        print('      MENÚ CLIENTES      ')
        print('--------------------------')
        print('1. Agregar cliente       ')
        print('2. Suspender cliente     ')
        print('3. Recuperar cliente     ')
        print('4. Consultas y reportes   ')
        print('5. Volver al menú principal   ')
        try:
          opcion = int(input('\nIngresa el número de la operación que deseas realizar: '))
          if opcion == 1:
              agregar_cliente()
          elif opcion == 2:
              suspender_cliente()
          elif opcion == 3:
              recuperar_cliente()
          elif opcion == 4:
              consultas_reportes_clientes()
          elif opcion == 5:
              return
          else:
              raise ValueError
        except ValueError:
              print('\n** OPCIÓN NO VÁLIDA. POR FAVOR, REGISTRE EL NÚMERO DE ALGUNA OPCIÓN MOSTRADA **')


def agregar_servicio():
    while True:
        print('\════════════════════════════')
        print('  AGREGAR UN NUEVO SERVICIO')
        print('════════════════════════════')
        print('\nNota: Escriba (0) si desea volver al menú clientes')
        while True:
            try:
                descripcion = input('\nIngrese el nombre del servicio: ')
                if descripcion == '0' :
                    print('\n** OPERACION CANCELADA, VOLVIENDO AL MENÚ DE SERVICIOS **')
                    return
                if descripcion.strip() == '':
                    print('\n** EL DATO NO PUEDE OMITIRSE, INGRESE UN DATO O (0) PARA VOLVER AL MENÚ DE SERVICIOS **')
                    continue
                if descripcion.isdigit():
                    print('\n** ERROR, INGRESE UN NOMBRE DE SERVICIO VÁLIDO O (0) PARA VOLVER AL MENÚ DE SERVICIOS**')
                    continue
                else:
                    break
            except ValueError as e:
                print(e)
        while True:
            try:
                costo = input('\nIngrese el costo del servicio: ')
                if costo.strip() == '':
                    print('\n** ERROR, EL DATO NO PUEDE OMITIRSE, INGRESE UN DATO VÁLIDO **')
                    continue
                costo = float(costo)
                if costo <= 0.0:
                    print('\n** EL COSTO DEL SERVICIO DEBE SER MAYOR A 0.0, INGRESE UN COSTO VÁLIDO **')
                    continue
                else:
                    break
            except ValueError:
                print('\n** DATO NO VÁLIDO, INGRESE UN COSTO **')
        try:
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                datos = (descripcion,costo)
                mi_cursor.execute('INSERT INTO servicios (descripcion, costo, estadoS) VALUES (?,?, 1)', datos)
                print('\n**Servicio registrado **')
                print(f'Clave asignada: {mi_cursor.lastrowid}')
        except Error as e:
            print(f'Se produjo el siguiente error: {e}')
        finally:
            conn.close()
        try:
            while True:
                agregar_servicio = input('\n¿Desea agregar otro servicio? (S)i (N)o: ')
                if agregar_servicio.lower() == 's':
                    break
                elif agregar_servicio.lower() == 'n':
                    print('\n** SERVICIO(S) REGISTRADO CORRECTAMENTE **')
                    raise StopIteration
                elif agregar_servicio.strip() == '':
                   print("ERROR, NO SE PUEDE OMITIR, INGRESA UN DATO VÁLIDO")
                else:
                    print('\n** DATO NO VÁLIDO, INGRESE (S)I Ó (N)O **')
                    continue
        except StopIteration:
            break

def suspender_servicio():
    while True:
        try:
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute('SELECT ClaveServicio, descripcion FROM servicios WHERE estadoS = 1')
                datos = mi_cursor.fetchall()
                if not datos:
                    print('\nNO HAY SERVICIOS DISPONIBLES.')
                else:
                    tabla = PrettyTable()
                    tabla.field_names = ["Clave Servicio", "Nombre"]
                    tabla.align["Clave Servicio"] = "r"
                    tabla.align["Nombre"] = "l"
                    for dato in datos:
                        tabla.add_row(dato)
                    print('\n       Servicios Suspendidos ')
                    print(tabla)
        except Exception as e:
            print(f'Se produjo el siguiente error: {e}')
        print('\n   SUSPENDER UN SERVICIO')
        print('\nNota: Escriba (0) si desea volver al menú principal')
        idservicio_suspender = input('Ingresa la clave del servicio a suspender: ')
        if idservicio_suspender == '0':
            return
        try:
            idservicio_suspender = int(idservicio_suspender)
        except ValueError:
            print('\n** ERROR, ingrese un número de servicio válido o escriba "0" para volver al menú principal **')
            continue
        try:
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute('SELECT ClaveServicio, descripcion, costo FROM servicios WHERE ClaveServicio = ? AND estadoS = 1', (idservicio_suspender,))
                servicio_detalle = mi_cursor.fetchone()
                if not servicio_detalle:
                    print('\n** ERROR, NO SE ENCONTRÓ UN SERVICIO CON ESA CLAVE O SE ENCUENTRA SUSPENDIDO **')
                    continue
                else:
                    tabla = PrettyTable()
                    tabla.field_names = ["Clave Servicio", "Nombre", "Costo"]
                    tabla.align["Clave Servicio"] = "r"
                    tabla.align["Nombre"] = "l"
                    tabla.align["Costo"] = "r"
                    tabla.add_row(servicio_detalle)
                    print('\n       Detalle del Servicio ')
                    print(tabla)
                while True:
                  confirmacion = input('\n¿Desea suspender este servicio? (S)i (N)o: ')
                  if confirmacion.lower() == 'n':
                    print("** SERVICIO NO SUSPENDIDO **")
                    menu_servicios()
                  elif confirmacion.strip() == '':
                     print("\n** ERROR, ESTE CAMPO NO SE PUEDE OMITIR, INGRESA UN DATO VÁLIDO **")
                  elif confirmacion.lower() == 's':
                    mi_cursor.execute('UPDATE servicios SET estadoS = 0 WHERE ClaveServicio = ?', (idservicio_suspender,))
                    conn.commit()
                    print('\n** Servicio suspendido correctamente **')
                    menu_servicios()
                  else:
                    print('\n** ERROR, entrada no válida. Por favor, ingrese "S" o "N" **')
        except sqlite3.Error as e:
          print(f'SE PRODUJO EL SIGUIENTE ERROR: {e}')

def recuperar_servicio():
    while True:
        try:
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute('SELECT ClaveServicio, descripcion, costo FROM servicios WHERE estadoS = 0')
                datos = mi_cursor.fetchall()
                if not datos:
                    print('\nNo hay servicios suspendidos.')
                else:
                    tabla_suspensos = PrettyTable()
                    tabla_suspensos.field_names = ["Clave Servicio", "Nombre", "Costo"]
                    for dato in datos:
                        tabla_suspensos.add_row(dato)
                    print('\n       Servicios Suspendidos ')
                    print(tabla_suspensos)
        except Exception as e:
            print(f'Se produjo el siguiente error: {e}')
        print('\n   RECUPERAR UN SERVICIO')
        print('\nNota: Escriba (0) si desea volver al menú principal')
        while True:
            idservicio_recuperar = input('Ingresa la clave del servicio a recuperar: ')
            if idservicio_recuperar == '0':
                return
            elif idservicio_recuperar.strip() == '':
                print("ERROR, ESTE CAMPO NO SE PUEDE OMITIR, INGRESE UN DATO VÁLIDO")
            elif idservicio_recuperar.isdigit():
                try:
                    with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute('SELECT ClaveServicio, descripcion, costo FROM servicios WHERE ClaveServicio = ? AND estadoS = 0', (idservicio_recuperar,))
                        servicio_detalle = mi_cursor.fetchone()
                        if not servicio_detalle:
                            print('\n** ERROR, NO SE ENCONTRÓ UN SERVICIO CON ESA CLAVE **')
                            continue
                        clave, descripcion, costo = servicio_detall
                        tabla_servicio = PrettyTable()
                        tabla_servicio.field_names = ["Clave Servicio", "Nombre", "Costo"]
                        tabla_servicio.add_row([clave, descripcion, costo])
                        print('\n       Detalle del Servicio ')
                        print(tabla_servicio)
                        while True:
                            confirmacion = input('\n¿Desea recuperar este servicio? (S)i (N)o: ').lower()
                            if confirmacion == 's':
                                mi_cursor.execute('UPDATE servicios SET estadoS = 1 WHERE ClaveServicio = ?', (idservicio_recuperar,))
                                conn.commit()
                                print('\n** Servicio recuperado correctamente **')
                                menu_servicios()
                            elif confirmacion == 'n':
                                print('\n** OPERACIÓN CANCELADA, VOLVIENDO AL MENÚ PRINCIPAL **')
                                break
                            elif confirmacion.strip() == '':
                                print("ERROR, ESTE CAMPO NO SE PUEDE OMITIR, INGRESE UN DATO VÁLIDO")
                            else:
                                print('\n** ERROR, entrada no válida. Por favor, ingrese "S" o "N" **')
                        break
                except sqlite3.Error as e:
                    print(f'Se produjo el siguiente error: {e}')
            else:
                print("** ERROR, INGRESA UN DATO VÁLIDO **")

def listado_servicios_registrados():
    while True:
        print('---------------------------------------')
        print('SUBMENÚ LISTADO DE SERVICIOS REGISTRADOS')
        print('---------------------------------------')
        print('1. Ordenado por clave')
        print('2. Ordenado por nombre de servicio')
        print('3. Volver al menú anterior')
        opcion = input('\nIngresa el número de la operación que deseas realizar: ')
        encabezados = ['Clave', 'Nombre de Servicio', 'Descripción', 'Costo']
        if opcion == '1':
            try:
                with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                  mi_cursor = conn.cursor()
                  mi_cursor.execute('SELECT ClaveServicio, descripcion, costo FROM servicios ORDER BY ClaveServicio')
                  datos = mi_cursor.fetchall()
                  if datos:
                    tabla = PrettyTable()
                    tabla.field_names = ["ClaveServicio", "Descripción", "Costo"]
                    for ClaveServicio, descripcion, costo in datos:
                      tabla.add_row([ClaveServicio, descripcion, costo])
                    print(tabla)
                    print("\nSERVICIOS REGISTRADOS")
                  else:
                    print("\n** NO HAY SERVICIOS REGISTRADOS **")
                  while True:
                    print('\n---------------------------------------')
                    print('           EXPORTAR REPORTE')
                    print('---------------------------------------')
                    print('1. Exportar reporte como archivo EXCEL')
                    print('2. Exportar reporte como archivo CSV')
                    print('3. Volver al menú de reportes')
                    exportar = input('\nIngresa el número de la operación que deseas realizar: ')
                    if exportar == '1':
                        fecha_reporte = datetime.now().strftime('%m_%d_%Y')
                        nombre_excel = f'ReporteServiciosPorClave_{fecha_reporte}.xlsx'
                        wb = Workbook()
                        hoja = wb.active
                        hoja.append(encabezados)
                        for dato in datos:
                            hoja.append(dato)
                        wb.save(nombre_excel)
                        print(f'\nInforme {nombre_excel} exportado correctamente')
                    elif exportar == '2':
                        fecha_reporte = datetime.now().strftime('%m_%d_%Y')
                        nombre_csv = f'ReporteServiciosPorClave_{fecha_reporte}.csv'
                        with open(nombre_csv, 'w', newline='') as reporte_csv:
                            grabador = csv.writer(reporte_csv)
                            grabador.writerow(encabezados)
                            grabador.writerows(datos)
                        print(f'\nInforme {nombre_csv} exportado correctamente')
                    elif exportar == '3':
                        break
                    elif exportar.strip() == '':
                      print("**ERROR, ESTE CAMPO NO SE PUEDE OMITIR**")
            except Exception as e:
                print(f"Error: {e}")
        elif opcion == '2':
            try:
                with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                  mi_cursor = conn.cursor()
                  mi_cursor.execute('SELECT ClaveServicio, descripcion, costo FROM servicios ORDER BY descripcion')
                  datos = mi_cursor.fetchall()
                  if datos:
                    tabla = PrettyTable()
                    tabla.field_names = ["ClaveServicio", "Descripción", "Costo"]
                    for ClaveServicio, descripcion, costo in datos:
                      tabla.add_row([ClaveServicio, descripcion, costo])
                    print(tabla)
                    print("\nSERVICIOS REGISTRADOS")
                  else:
                    print("\n** NO HAY SERVICIOS REGISTRADOS **")
                while True:
                  print('\n---------------------------------------')
                  print('           EXPORTAR REPORTE')
                  print('---------------------------------------')
                  print('1. Exportar reporte como archivo EXCEL')
                  print('2. Exportar reporte como archivo CSV')
                  print('3. Volver al menú de reportes')
                  exportar = input('\nIngresa el número de la operación que deseas realizar: ')
                  if exportar == '1':
                        fecha_reporte = datetime.now().strftime('%m_%d_%Y')
                        nombre_excel = f'ReporteServiciosPorNombre_{fecha_reporte}.xlsx'
                        wb = Workbook()
                        hoja = wb.active
                        hoja.append(encabezados)
                        for dato in datos:
                            hoja.append(dato)
                        wb.save(nombre_excel)
                        print(f'\nInforme {nombre_excel} exportado correctamente')
                  elif exportar == '2':
                        fecha_reporte = datetime.now().strftime('%m_%d_%Y')
                        nombre_csv = f'ReporteServiciosPorNombre_{fecha_reporte}.csv'
                        with open(nombre_csv, 'w', newline='') as reporte_csv:
                            grabador = csv.writer(reporte_csv)
                            grabador.writerow(encabezados)
                            grabador.writerows(datos)
                        print(f'\nInforme {nombre_csv} exportado correctamente')
                  elif exportar == '3':
                        break
                  elif exportar.strip() == '':
                    print("** ERROR, ESTE CAMPO NO SE PUEDE OMITIR **")
                  else:
                    print("** ERROR, INGRESE UN DATO VÁLIDO **")
            except Exception as e:
                print(f"Error: {e}")
        elif opcion == '3':
            break
        elif opcion.strip() == '':
          print("** ERROR, ESTE CAMPO NO SE PUEDE OMITIR **")

def consultas_reportes_servicios():
    while True:
        print('\n  ═══════════════════════════════════')
        print('   CONSULTAS Y REPORTES DE SERVICIOS')
        print('  ═══════════════════════════════════')
        print('1. Listado de servicios registrados')
        print('2. Búsqueda por clave')
        print('3. Búsqueda por nombre')
        print('4. Volver al menú servicios')
        consulta = input('\nIngresa el número de la operación que desea realizar: ')
        if consulta == '1':
          listado_servicios_registrados()
        elif consulta == '2':
           while True:
              try:
                with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                  mi_cursor = conn.cursor()
                  mi_cursor.execute('SELECT ClaveServicio, descripcion FROM servicios ORDER BY ClaveServicio')
                  datos_servicios = mi_cursor.fetchall()
                  tabla_servicios = PrettyTable()
                  tabla_servicios.field_names = ['CLAVE', 'DESCRIPCION']
                  for servicio in datos_servicios:
                    tabla_servicios.add_row(servicio)
                  print(tabla_servicios)
                clave_consultar = input('\nIngrese la clave del servicio a consultar (o escriba "0" para volver): ')
                if clave_consultar.lower() == '0':
                    break
                try:
                    clave_consultar = int(clave_consultar)
                except ValueError:
                    print('\n** ERROR, INGRESE UN NÚMERO DE SERVICIO VÁLIDO O ESCRIBA "0" PARA REGRESAR AL MENU ANTERIOR **')
                mi_cursor.execute('SELECT ClaveServicio, descripcion, costo FROM servicios WHERE ClaveServicio = ?', (clave_consultar,))
                datos_servicio_consultado = mi_cursor.fetchone()
                if not datos_servicio_consultado:
                    print(f'\n** NO SE ENCONTRÓ UN REGISTRO ASOCIADO A LA CLAVE {clave_consultar} **')
                else:
                    tabla_servicio_consultado = PrettyTable()
                    tabla_servicio_consultado.field_names = ['CLAVE', 'DESCRIPCION', 'COSTO']
                    tabla_servicio_consultado.add_row(datos_servicio_consultado)
                    print(f'\n    Detalle del Servicio Consultado')
                    print(tabla_servicio_consultado)
                    break
              except Exception as e:
                print(f'Se produjo el siguiente error: {e}')
        elif consulta == '3':
                while True:
                    print('\nNota: Escriba (0) si desea volver a consultas y reportes de servicios')
                    try:
                        nombre_consultar = input('\nIngrese el nombre del servicio a consultar: ')
                        if nombre_consultar == 0:
                            break
                        elif nombre_consultar.strip() == '':
                          print("ERROR, ESTE CAMPO NO SE PUEDE OMITIR, INGRESE UN DATO VÁLIDO")
                        elif nombre_consultar.isdigit():
                          print("**ERROR, INGRESE UN NOMBRE DE SERVICIO VÁLIDO **")
                        else:
                            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                                mi_cursor = conn.cursor()
                                datos = {'nombre': nombre_consultar.lower()}
                                mi_cursor.execute('SELECT ClaveServicio, descripcion, costo FROM servicios WHERE LOWER (descripcion) = :nombre', datos)
                                datos1 = mi_cursor.fetchall()
                                tabla = PrettyTable()
                                tabla.field_names = ['CLAVE', 'DESCRIPCION', 'COSTO']
                                if datos1:
                                    print(f'\n    Servicio Encontrado')
                                    for dato in datos1:
                                        tabla.add_row(dato)
                                    print(tabla)
                                    break
                                else:
                                    print(f'\n** NO SE ENCONTRÓ UN REGISTRO ASOCIADO AL {nombre_consultar}')
                    except Exception:
                        print(f'\n** DATO NO VÁLIDO. INGRESE EL NOMBRE DE ALGÚN SERVICIO **')
        elif consulta == '4':
                return
        elif consulta.strip() == ' ':
          print("** ERROR, NO SE PUEDE OMITIR ESTE CAMPO **")
        else:
          print("** ERROR, INGRESE UN DATO VÁLIDO **")

def menu_servicios():
    while True:
        print('\n---------------------------------------')
        print('  BIENVENIDO AL MENU DE SERVICIOS   ')
        print('---------------------------------------')
        print('1. Agregar un servicio')
        print('2. Suspender un servicio')
        print('3. Recuperar un servicio')
        print('4. Consultas y reportes')
        print('5. Regresar al menú principal')
        opcion = input('\nIngrese el número de la opcion a la que desea ingresar: ')
        if opcion.isdigit():
            opcion = int(opcion)
            if opcion == 1:
                agregar_servicio()
            elif opcion == 2:
                suspender_servicio()
            elif opcion == 3:
                recuperar_servicio()
            elif opcion == 4:
                consultas_reportes_servicios()
            elif opcion == 5:
                while True:
                   respuesta = input("¿Desea salir? (S/N) ")
                   if respuesta.upper() == 'S':
                      menu_principal()
                   elif respuesta.upper() == 'N':
                       menu_servicios()
                   elif respuesta.strip() == '':
                      print("No se puede omitir")
                   elif respuesta.isdigit():
                      print("Ingrese una opción ")
                   else:
                       print("**Ingrese una opción válida")
            else:
                print("Opción no válida. Por favor, elige una opción válida.")
        else:
            print('\nOpción no válida. Por favor, elige una opción válida.')

def obtener_servicios_mas_prestados():
    while True:
        try:
            print('\n-------------------------------------')
            print('      SERVICIOS MÁS PRESTADOS')
            print('--------------------------------------')
            cantidad_servicios = int(input('\nIngrese la cantidad de servicios más prestados a identificar o ingrese "0" para regresar al menú de estadisticas: '))
            if cantidad_servicios == 0:
                menu_estadisticas()
                return
            elif cantidad_servicios < 1:
                print('\n** LA CANTIDAD DE SERVICIOS A OBSERVAR DEBE SER AL MENOS 1. **')
                continue
            while True:
                fecha_inicial_str = input('\nIngrese la fecha inicial del período a reportar (DD/MM/YYYY): ')
                try:
                    fecha_inicial = datetime.strptime(fecha_inicial_str, "%d/%m/%Y").date()
                    if fecha_inicial > datetime.now().date():
                        print('\n** LA FECHA INICIAL NO PUEDE SER POSTERIOR A LA FECHA ACTUAL. **')
                        continue
                    break
                except ValueError:
                    print('\n** ERROR, FORMATO DE FECHA INCORRECTO. INTENTE NUEVAMENTE **')
            while True:
                fecha_final_str = input('\nIngrese la fecha final del período a reportar (DD/MM/YYYY): ')
                try:
                    fecha_final = datetime.strptime(fecha_final_str, "%d/%m/%Y").date()
                    if fecha_final > datetime.now().date():
                        print('\n** LA FECHA FINAL NO PUEDE SER POSTERIOR A LA FECHA ACTUAL. **')
                        continue
                    if fecha_final < fecha_inicial:
                        print('\n** LA FECHA FINAL DEBE SER MAYOR O IGUAL A LA FECHA INICIAL. **')
                        continue
                    break
                except ValueError:
                    print('\n** ERROR, FORMATO DE FECHA INCORRECTO. INTENTE NUEVAMENTE **')
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute('''
                    SELECT s.descripcion, COUNT(d.ClaveServicio) as cantidad_prestada
                    FROM servicios s
                    JOIN detalle d ON s.ClaveServicio = d.ClaveServicio
                    JOIN notas n ON d.folio = n.folio
                    WHERE n.fecha BETWEEN ? AND ?
                    GROUP BY s.descripcion
                    ORDER BY cantidad_prestada DESC
                    LIMIT ?;
                ''', (fecha_inicial, fecha_final, cantidad_servicios))
                datos = mi_cursor.fetchall()
                if not datos:
                    print('\nNo hay servicios prestados en el período especificado.')
                else:
                    tabla = PrettyTable()
                    tabla.field_names = ["Servicio", "Cantidad Prestada"]
                    for servicio, cantidad_prestada in datos:
                        tabla.add_row([servicio, cantidad_prestada])
                    print('\n       Servicios Más Prestados ')
                    print(tabla)
                    while True:
                        print('\n---------------------------------------')
                        print('           EXPORTAR REPORTE')
                        print('---------------------------------------')
                        print('1. Exportar reporte como archivo EXCEL')
                        print('2. Exportar reporte como archivo CSV')
                        print('3. Volver al menú de Estadísticas')
                        exportar = int(input('\nIngresa el número de la operación que deseas realizar: '))
                        if exportar == 1:
                            fecha_reporte = datetime.now().strftime('%d_%m_%Y')
                            nombre_excel = f'ReporteServiciosMasPrestados_{fecha_inicial.strftime("%d_%m_%Y")}_{fecha_final.strftime("%d_%m_%Y")}.xlsx'
                            wb = Workbook()
                            hoja = wb.active
                            hoja.append(["Servicio", "Cantidad Prestada"])
                            for servicio, cantidad_prestada in datos:
                                hoja.append([servicio, cantidad_prestada])
                            wb.save(nombre_excel)
                            print(f'\nInforme {nombre_excel} exportado correctamente')
                            menu_estadisticas()
                        elif exportar == 2:
                            fecha_reporte = datetime.now().strftime('%d_%m_%Y')
                            nombre_csv = f'ReporteServiciosMasPrestados_{fecha_inicial.strftime("%d_%m_%Y")}_{fecha_final.strftime("%d_%m_%Y")}.csv'
                            with open(nombre_csv, 'w', newline='') as reporte_csv:
                                grabador = csv.writer(reporte_csv)
                                grabador.writerow(["Servicio", "Cantidad Prestada"])
                                grabador.writerows(datos)
                            print(f'\nInforme {nombre_csv} exportado correctamente')
                            menu_estadisticas()
                        elif exportar == 3:
                            print('\nVolviendo al Menú de Estadísticas.')
                            menu_estadisticas()
                        else:
                            print('\n** ERROR, OPCION NO VÁLIDA, INTENTE NUEVAMENTE.')
        except ValueError:
            print('\n** DEBE INGRESAR UN NÚMERO ENTERO VÁLIDO. **')
        except Exception as e:
            print(f'Se produjo el siguiente error: {e}')
            
def clientes_mas_notas():
    while True:
        try:
            cantidad_clientes = int(input('\nIngrese la cantidad de clientes con más notas a identificar o ingrese "0" para regresar al menú anterior: '))
            if cantidad_clientes == 0:
                menu_estadisticas()
                return                  
            elif cantidad_clientes < 1:
                print('\n** LA CANTIDAD DE CLIENTES A OBSERVAR DEBE SER AL MENOS 1. **')
                continue            
            while True:
                fecha_inicial_str = input('\nIngrese la fecha inicial del período a reportar (DD/MM/YYYY): ')
                try:
                    fecha_inicial = datetime.strptime(fecha_inicial_str, "%d/%m/%Y").date()

                    if fecha_inicial > datetime.now().date():
                        print('\n** LA FECHA INICIAL NO PUEDE SER MAYOR A LA FECHA ACTUAL. **')
                        continue
                    break
                except ValueError:
                    print('\n** ERROR, FORMATO DE FECHA INCORRECTO. INTENTE NUEVAMENTE **')
            while True:
                fecha_final_str = input('\nIngrese la fecha final del período a reportar (DD/MM/YYYY): ')
                try:
                    fecha_final = datetime.strptime(fecha_final_str, "%d/%m/%Y").date()
                    if fecha_final > datetime.now().date():
                        print('\n** LA FECHA FINAL NO PUEDE SER POSTERIOR A LA FECHA ACTUAL. **')
                        continue
                    if fecha_final < fecha_inicial:
                        print('\n** LA FECHA FINAL DEBE SER MAYOR O IGUAL A LA FECHA INICIAL. **')
                        continue
                    break
                except ValueError:
                    print('\n** ERROR, FORMATO DE FECHA INCORRECTO. INTENTE NUEVAMENTE **')

            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute('''
                SELECT clientes.ClaveCliente, clientes.nombre, COUNT(notas.folio) as cantidad_notas
                FROM clientes
                LEFT JOIN notas ON clientes.ClaveCliente = notas.ClaveCliente
                WHERE notas.fecha BETWEEN ? AND ?
                GROUP BY clientes.ClaveCliente
                ORDER BY cantidad_notas DESC
                LIMIT ?;
                ''', (fecha_inicial, fecha_final, cantidad_clientes))
                clientes_con_mas_notas = mi_cursor.fetchall()
                if clientes_con_mas_notas:
                    encabezados = ['Clave Cliente', 'Nombre', 'Cantidad de Notas']
                    print('\n---------------------------------------')
                    print('     CLIENTES CON MÁS NOTAS')
                    print('---------------------------------------')
                    tabla = PrettyTable()
                    tabla.field_names = encabezados
                    for cliente in clientes_con_mas_notas:
                        tabla.add_row(cliente)
                    print(tabla)
                    print('\n---------------------------------------')
                    print('           EXPORTAR REPORTE')
                    print('---------------------------------------')
                    print('1. Exportar reporte como archivo EXCEL')
                    print('2. Exportar reporte como archivo CSV')
                    print('3. Volver al menú de Estadísticas')
                    exportar = int(input('\nIngresa el número de la operación que deseas realizar: '))
                    if exportar == 1:
                        fecha_reporte = datetime.now().strftime('%d_%m_%Y')
                        nombre_excel = f'ReporteClientesConMasNotas_{fecha_inicial.strftime("%d_%m_%Y")}_{fecha_final.strftime("%d_%m_%Y")}.xlsx'
                        wb = Workbook()
                        hoja = wb.active
                        hoja.append(encabezados)
                        hoja.append(['Fecha Inicial', fecha_inicial_str, 'Fecha Final', fecha_final_str])
                        for cliente in clientes_con_mas_notas:
                            hoja.append(cliente)
                        wb.save(nombre_excel)
                        print(f'\nInforme {nombre_excel} exportado correctamente')
                        menu_estadisticas()
                    elif exportar == 2:
                        fecha_reporte = datetime.now().strftime('%d_%m_%Y')
                        nombre_csv = f'ReporteClientesConMasNotas_{fecha_inicial.strftime("%d_%m_%Y")}_{fecha_final.strftime("%d_%m_%Y")}.csv'
                        with open(nombre_csv, 'w', newline='') as reporte_csv:
                            grabador = csv.writer(reporte_csv)
                            grabador.writerow(encabezados)
                            grabador.writerow(['Fecha Inicial', fecha_inicial_str, 'Fecha Final', fecha_final_str])
                            grabador.writerows(clientes_con_mas_notas)
                        print(f'\nInforme {nombre_csv} exportado correctamente')
                        menu_estadisticas()
                    elif exportar == 3:
                        print('\nVolviendo al menú de Estadísticas.')
                        menu_estadisticas()
                    else:
                        print('\n** ERROR, OPCION NO VÁLIDA, INTENTE NUEVAMENTE.')
                else:
                    print('\n** NO HAY CLIENTES CON NOTAS EN EL PERIODO SELECCIONADO **.')
        except ValueError:
            print('\n** DEBE INGRESAR UN NÚMERO ENTERO VÁLIDO. **')
        except Exception as e:
            print(f'Error: {e}')

def promedio_montos_notas():
   while True:
        try:
            print('\n   ----------------------------------------')
            print('      PROMEDIO DE LOS MONTOS DE LAS NOTAS  ')
            print('   ----------------------------------------')            
            while True:                
                fecha_inicial_str = input('\nIngrese la fecha inicial del período a reportar (DD/MM/YYYY): ')
                try:
                    fecha_inicial = datetime.strptime(fecha_inicial_str, "%d/%m/%Y").date()
                    if fecha_inicial > datetime.now().date():
                        print('\n** LA FECHA INICIAL NO PUEDE SER POSTERIOR A LA FECHA ACTUAL. **')
                        continue
                    break
                except ValueError:
                      print('\n** ERROR, FORMATO DE FECHA INCORRECTO. INTENTE NUEVAMENTE **')
            while True:
                  fecha_final_str = input('\nIngrese la fecha final del período a reportar (DD/MM/YYYY): ')
                  try:
                     fecha_final = datetime.strptime(fecha_final_str, "%d/%m/%Y").date()
                     if fecha_final > datetime.now().date():
                        print('\n** LA FECHA FINAL NO PUEDE SER POSTERIOR A LA FECHA ACTUAL. **')
                        continue
                     if fecha_final < fecha_inicial:
                        print('\n** LA FECHA FINAL DEBE SER MAYOR O IGUAL A LA FECHA INICIAL. **')
                        continue
                     break
                  except ValueError:
                    print('\n** ERROR, FORMATO DE FECHA INCORRECTO. INTENTE NUEVAMENTE **')                
            with sqlite3.connect('C:/Users/betyh/Downloads/tallermecanico.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute('''
                    SELECT AVG(monto) as promedio_monto
                    FROM notas
                   WHERE fecha BETWEEN ? AND ?
                  ''', (fecha_inicial, fecha_final))                                  
                promedio = mi_cursor.fetchall()[0][0]
                print(f"{'-' * 80}\n")
                print(f"   El monto promedio de las notas para el período seleccionado es: {promedio:.2f}\n")
                print(f"{'-' * 80}\n")
                menu_estadisticas()
        except Exception as e:
           print(f'Se produjo el siguiente error: {e}')
           
def menu_estadisticas():
    while True:
        print('---------------------------------------')
        print('     MENÚ DE ESTADÍSTICAS         ')
        print('---------------------------------------')
        print('1. Servicios más prestados')
        print('2. Clientes con más notas')
        print('3. Promedio de montos de las notas')
        print('4. Salir')
        opcion_estadisticas = input('Ingrese el número de la opción deseada: ')
        if opcion_estadisticas.isdigit():
            opcion_estadisticas = int(opcion_estadisticas)
            if opcion_estadisticas == 1:
                obtener_servicios_mas_prestados()
            elif opcion_estadisticas == 2:
                clientes_mas_notas()
            elif opcion_estadisticas == 3:
                promedio_montos_notas()
            elif opcion_estadisticas == 4:
                menu_principal()
            else:
                print("** ERROR, INGRESE UNA OPCIÓN VÁLIDA **")
        else:
            print('\n** ERROR, INGRESE UNA OPCIÓN VÁLIDA **')
            
def menu_principal():
    while True:
        print('---------------------------------------')
        print('  BIENVENIDO AL MENU PRINCIPAL   ')
        print('---------------------------------------')
        print('1. Notas')
        print('2. Clientes')
        print('3. Servicios')
        print('4. Estadísticas')
        print('5. Salir')
        opcion = input('Ingrese el número del menú al que desea ingresar: ')
        if opcion.isdigit():
            opcion = int(opcion)
            if opcion == 1:
                print("Entraste al menú de notas")
                menu_notas()
            elif opcion == 2:
                menu_clientes()
            elif opcion == 3:
                print("Entraste a Servicios")
                menu_servicios()
            elif opcion == 4:
                print("Entraste a Estadísticas")
                menu_estadisticas()
            elif opcion == 5:
                respuesta = input("¿Desea salir? (S/N) ")
                if respuesta.lower() == 's':
                    break
            else:
                print("Opción no válida. Por favor, elige una opción válida.")
        else:
            print('\nOpción no válida. Por favor, elige una opción válida.')

menu_principal()



